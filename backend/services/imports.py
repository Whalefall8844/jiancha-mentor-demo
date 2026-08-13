from __future__ import annotations

import csv
import hashlib
import json
import re
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

from ..repositories.catalog import (
    create_project,
    create_site,
    get_project,
    get_project_by_code,
    get_site,
    get_site_by_project_and_code,
    list_subject_codes,
    save_subject_codes,
    update_project,
    update_site,
)
from ..database import get_connection, transaction
from ..repositories.visits import create_audit_event


PROJECT_ALIASES = ("project_code", "项目编号", "项目编码", "研究编号", "study_id")
SITE_ALIASES = ("site_code", "中心编号", "中心编码", "site_id")
SUBJECT_ALIASES = ("subject_code", "受试者编号", "受试者代码", "subject_id")


def _normalise_key(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def _string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _value(row: dict[str, str], *aliases: str) -> str:
    normalized = {_normalise_key(key): _string(value) for key, value in row.items()}
    for alias in aliases:
        value = normalized.get(_normalise_key(alias), "")
        if value:
            return value
    return ""


def _site_code(row: dict[str, str]) -> str:
    code = _value(row, *SITE_ALIASES)
    return code.zfill(3) if code.isdecimal() and len(code) < 3 else code


def _rows_from_xlsx(content: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = [_string(value) for value in next(rows, ())]
    if not headers or not any(headers):
        return []
    return [
        {headers[index]: _string(value) for index, value in enumerate(values) if index < len(headers) and headers[index]}
        for values in rows
        if any(_string(value) for value in values)
    ]


def _rows_from_csv(content: bytes) -> list[dict[str, str]]:
    decoded = ""
    for encoding in ("utf-8-sig", "gb18030", "gbk"):
        try:
            decoded = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if not decoded:
        raise ValueError("无法识别 CSV 文件编码")
    try:
        dialect = csv.Sniffer().sniff(decoded[:2048], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    return [
        {str(key or "").strip(): _string(value) for key, value in row.items() if key}
        for row in csv.DictReader(StringIO(decoded), dialect=dialect)
        if any(_string(value) for value in row.values())
    ]


def read_rows(*, file_name: str, content: bytes) -> list[dict[str, str]]:
    suffix = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""
    if suffix in {"xlsx", "xlsm"}:
        return _rows_from_xlsx(content)
    if suffix in {"csv", "txt", "tsv"}:
        return _rows_from_csv(content)
    raise ValueError("仅支持 .xlsx、.xlsm、.csv 或 .tsv 文件")


def _project_for_row(row: dict[str, str], default_project_id: str) -> dict[str, Any] | None:
    project_code = _value(row, *PROJECT_ALIASES)
    if project_code:
        return get_project_by_code(project_code)
    return get_project(default_project_id) if default_project_id else None


def _site_for_row(row: dict[str, str], *, project_id: str, default_site_id: str) -> dict[str, Any] | None:
    site_code = _site_code(row)
    if site_code:
        return get_site_by_project_and_code(project_id, site_code)
    return get_site(default_site_id) if default_site_id else None


def _audit(project_id: str, entity_type: str, entity_id: str, actor_name: str, detail: dict[str, Any]) -> None:
    create_audit_event(
        project_id=project_id,
        visit_id=None,
        entity_type=entity_type,
        entity_id=entity_id,
        action="imported",
        actor_name=actor_name,
        detail=detail,
    )


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M")


def _batch_row(row) -> dict[str, Any]:
    data = dict(row)
    data["operation"] = json.loads(data.pop("operation_json") or "{}")
    data["source_row"] = json.loads(data.pop("source_row_json") or "{}")
    return data


def _summary_from_rows(rows: list[dict[str, Any]]) -> dict[str, int]:
    summary = {"total": len(rows), "created": 0, "updated": 0, "valid": 0, "skipped": 0}
    for row in rows:
        if row.get("error_message"):
            summary["skipped"] += 1
        elif row.get("action") == "create":
            summary["created"] += 1
            summary["valid"] += 1
        elif row.get("action") == "update":
            summary["updated"] += 1
            summary["valid"] += 1
        else:
            summary["skipped"] += 1
    return summary


def _normalize_column_mapping(value: dict[str, Any] | None) -> dict[str, str]:
    if not isinstance(value, dict):
        return {}
    return {
        _string(key): _string(source)
        for key, source in value.items()
        if _string(key) and _string(source)
    }


def _apply_column_mapping(row: dict[str, str], column_mapping: dict[str, str]) -> dict[str, str]:
    if not column_mapping:
        return row
    available = {_normalise_key(key): _string(value) for key, value in row.items()}
    mapped = dict(row)
    for target_field, source_header in column_mapping.items():
        value = available.get(_normalise_key(source_header), "")
        if value:
            mapped[target_field] = value
    return mapped


def _preflight_project(row: dict[str, str]) -> dict[str, Any]:
    code = _value(row, *PROJECT_ALIASES)
    name = _value(row, "project_name", "项目名称", "研究名称", "study_name")
    if not code or not name:
        raise ValueError("缺少项目编号或项目名称")
    metadata = {
        "approval_number": _value(row, "approval_number", "nmpa批件号", "批件号", "批准文号"),
        "sop_version": _value(row, "sop_version", "sop版本", "监查sop版本"),
    }
    existing = get_project_by_code(code)
    if existing:
        payload = {
            "project_id": existing["id"],
            "name": name,
            "sponsor": _value(row, "sponsor", "申办方") or existing.get("sponsor", ""),
            "metadata": {**existing.get("metadata", {}), **{key: value for key, value in metadata.items() if value}},
        }
        return {"action": "update", "entity_type": "project", "dedupe_key": code.upper(), "payload": payload}
    return {
        "action": "create",
        "entity_type": "project",
        "dedupe_key": code.upper(),
        "payload": {
            "code": code,
            "name": name,
            "sponsor": _value(row, "sponsor", "申办方"),
            "metadata": {key: value for key, value in metadata.items() if value},
        },
    }


def _preflight_site(row: dict[str, str], default_project_id: str) -> dict[str, Any]:
    project = _project_for_row(row, default_project_id)
    code = _site_code(row)
    name = _value(row, "site_name", "中心名称", "研究中心名称")
    if project is None:
        raise ValueError("未找到对应项目；请提供项目编号或选择当前项目")
    if not code or not name:
        raise ValueError("缺少中心编号或中心名称")
    patch = {
        "name": name,
        "pi_name": _value(row, "pi_name", "中心pi", "主要研究者", "pi"),
        "ethics_date": _value(row, "ethics_date", "伦理日期", "伦理批准日期"),
        "protocol_version": _value(row, "protocol_version", "方案版本", "protocol"),
        "icf_version": _value(row, "icf_version", "知情同意书版本", "icf版本"),
    }
    existing = get_site_by_project_and_code(project["id"], code)
    if existing:
        payload = {key: value or existing.get(key, "") for key, value in patch.items()}
        payload["site_id"] = existing["id"]
        return {
            "action": "update",
            "entity_type": "site",
            "dedupe_key": f"{project['id']}:{code.upper()}",
            "payload": payload,
        }
    return {
        "action": "create",
        "entity_type": "site",
        "dedupe_key": f"{project['id']}:{code.upper()}",
        "payload": {"project_id": project["id"], "code": code, **patch},
    }


def _preflight_subject(row: dict[str, str], default_project_id: str, default_site_id: str) -> dict[str, Any]:
    project = _project_for_row(row, default_project_id)
    if project is None:
        raise ValueError("未找到对应项目；请提供项目编号或选择当前项目")
    site = _site_for_row(row, project_id=project["id"], default_site_id=default_site_id)
    subject_code = _value(row, *SUBJECT_ALIASES).upper()
    if site is None:
        raise ValueError("未找到对应中心；请提供中心编号或选择当前中心")
    if not subject_code:
        raise ValueError("缺少受试者编号")
    before = {item["code"] for item in list_subject_codes(site["id"])}
    return {
        "action": "update" if subject_code in before else "create",
        "entity_type": "subject_code",
        "dedupe_key": f"{site['id']}:{subject_code}",
        "payload": {
            "project_id": project["id"],
            "site_id": site["id"],
            "subject_code": subject_code,
            "enrollment_status": _value(row, "enrollment_status", "入组状态", "状态") or "screening",
        },
    }


def _preflight_operation(
    scope: str,
    row: dict[str, str],
    default_project_id: str,
    default_site_id: str,
) -> dict[str, Any]:
    if scope == "projects":
        return _preflight_project(row)
    if scope == "sites":
        return _preflight_site(row, default_project_id)
    return _preflight_subject(row, default_project_id, default_site_id)


def get_import_batch(batch_id: str) -> dict[str, Any] | None:
    with get_connection() as connection:
        batch = connection.execute("SELECT * FROM import_batches WHERE id = ?", (batch_id,)).fetchone()
        rows = connection.execute(
            "SELECT * FROM import_batch_rows WHERE batch_id = ? ORDER BY row_number",
            (batch_id,),
        ).fetchall()
    if batch is None:
        return None
    data = dict(batch)
    preview_summary = json.loads(data.pop("preview_summary_json") or "{}")
    source = preview_summary.pop("source", {})
    profile = preview_summary.pop("import_profile", {})
    data["preview_summary"] = preview_summary
    if isinstance(source, dict):
        data["source_system"] = str(source.get("system") or "")
        data["source_reference"] = str(source.get("reference") or "")
        data["source_exported_at"] = str(source.get("exported_at") or "")
    if isinstance(profile, dict):
        data["import_profile_id"] = str(profile.get("id") or "")
        data["import_profile_name"] = str(profile.get("name") or "")
    data["committed_summary"] = json.loads(data.pop("committed_summary_json") or "{}")
    data["rows"] = [_batch_row(row) for row in rows]
    return data


def preview_master_data_import(
    *,
    scope: str,
    file_name: str,
    content: bytes,
    default_project_id: str = "",
    default_site_id: str = "",
    actor_name: str = "演示管理员",
    source_system: str = "",
    source_reference: str = "",
    source_exported_at: str = "",
    import_profile_id: str = "",
    import_profile_name: str = "",
    column_mapping: dict[str, Any] | None = None,
) -> dict[str, Any]:
    normalized_scope = scope.strip().lower()
    if normalized_scope not in {"projects", "sites", "subjects"}:
        raise ValueError("导入类型仅支持 projects、sites 或 subjects")
    rows = read_rows(file_name=file_name, content=content)
    normalized_mapping = _normalize_column_mapping(column_mapping)
    preview_rows: list[dict[str, Any]] = []
    dedupe_rows: dict[str, int] = {}
    for offset, row in enumerate(rows, start=2):
        try:
            operation = _preflight_operation(
                normalized_scope,
                _apply_column_mapping(row, normalized_mapping),
                default_project_id,
                default_site_id,
            )
            dedupe_key = operation.pop("dedupe_key")
            if dedupe_key in dedupe_rows:
                raise ValueError(f"本批次与第 {dedupe_rows[dedupe_key]} 行主键重复")
            dedupe_rows[dedupe_key] = offset
            preview_rows.append(
                {
                    "row_number": offset,
                    "action": operation["action"],
                    "entity_type": operation["entity_type"],
                    "operation": operation["payload"],
                    "source_row": row,
                    "error_message": "",
                }
            )
        except ValueError as exc:
            preview_rows.append(
                {
                    "row_number": offset,
                    "action": "skip",
                    "entity_type": normalized_scope,
                    "operation": {},
                    "source_row": row,
                    "error_message": str(exc),
                }
            )
    summary = _summary_from_rows(preview_rows)
    source = {
        "system": source_system.strip(),
        "reference": source_reference.strip(),
        "exported_at": source_exported_at.strip(),
    }
    source = {key: value for key, value in source.items() if value}
    if source:
        summary["source"] = source
    import_profile = {
        "id": import_profile_id.strip(),
        "name": import_profile_name.strip(),
        "column_mapping": normalized_mapping,
    }
    import_profile = {key: value for key, value in import_profile.items() if value}
    if import_profile:
        summary["import_profile"] = import_profile
    batch_id = uuid4().hex
    timestamp = _now()
    with transaction() as connection:
        connection.execute(
            """
            INSERT INTO import_batches (
                id, scope, file_name, content_hash, default_project_id, default_site_id,
                actor_name, status, preview_summary_json, committed_summary_json, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 'previewed', ?, '{}', ?)
            """,
            (
                batch_id,
                normalized_scope,
                file_name,
                hashlib.sha256(content).hexdigest(),
                default_project_id,
                default_site_id,
                actor_name,
                json.dumps(summary, ensure_ascii=False),
                timestamp,
            ),
        )
        for row in preview_rows:
            connection.execute(
                """
                INSERT INTO import_batch_rows (id, batch_id, row_number, action, entity_type, operation_json, source_row_json, error_message, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    uuid4().hex,
                    batch_id,
                    row["row_number"],
                    row["action"],
                    row["entity_type"],
                    json.dumps(row["operation"], ensure_ascii=False),
                    json.dumps(row["source_row"], ensure_ascii=False),
                    row["error_message"],
                    timestamp,
                ),
            )
    return get_import_batch(batch_id) or {}


def _int_value(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _quality_rate(valid_rows: int, total_rows: int) -> float:
    if total_rows <= 0:
        return 0
    return round(valid_rows * 100 / total_rows, 1)


def get_project_import_quality(*, project_id: str, recent_limit: int = 12) -> dict[str, Any]:
    with get_connection() as connection:
        rows = connection.execute(
            """
            SELECT * FROM import_batches
            WHERE default_project_id = ?
            ORDER BY created_at DESC, id DESC
            """,
            (project_id,),
        ).fetchall()

    aggregate = {
        "total_batches": 0,
        "committed_batches": 0,
        "previewed_batches": 0,
        "total_rows": 0,
        "valid_rows": 0,
        "skipped_rows": 0,
        "source_traced_batches": 0,
    }
    scopes: dict[str, dict[str, Any]] = {}
    batches: list[dict[str, Any]] = []
    last_imported_at = ""
    for row in rows:
        data = dict(row)
        preview_summary = json.loads(data.get("preview_summary_json") or "{}")
        source = preview_summary.get("source") if isinstance(preview_summary.get("source"), dict) else {}
        profile = preview_summary.get("import_profile") if isinstance(preview_summary.get("import_profile"), dict) else {}
        total_rows = _int_value(preview_summary.get("total"))
        valid_rows = _int_value(preview_summary.get("valid"))
        skipped_rows = _int_value(preview_summary.get("skipped"))
        scope = str(data.get("scope") or "")

        aggregate["total_batches"] += 1
        aggregate["total_rows"] += total_rows
        aggregate["valid_rows"] += valid_rows
        aggregate["skipped_rows"] += skipped_rows
        if data.get("status") == "committed":
            aggregate["committed_batches"] += 1
            if not last_imported_at and data.get("committed_at"):
                last_imported_at = str(data["committed_at"])
        else:
            aggregate["previewed_batches"] += 1
        if source:
            aggregate["source_traced_batches"] += 1

        scope_summary = scopes.setdefault(scope, {"scope": scope, "batch_count": 0, "total_rows": 0, "valid_rows": 0, "skipped_rows": 0})
        scope_summary["batch_count"] += 1
        scope_summary["total_rows"] += total_rows
        scope_summary["valid_rows"] += valid_rows
        scope_summary["skipped_rows"] += skipped_rows

        if len(batches) < recent_limit:
            batches.append(
                {
                    "id": str(data["id"]),
                    "scope": scope,
                    "file_name": str(data["file_name"]),
                    "status": str(data["status"]),
                    "created_at": str(data["created_at"]),
                    "committed_at": str(data.get("committed_at") or ""),
                    "total_rows": total_rows,
                    "valid_rows": valid_rows,
                    "skipped_rows": skipped_rows,
                    "quality_rate": _quality_rate(valid_rows, total_rows),
                    "source_system": str(source.get("system") or ""),
                    "source_reference": str(source.get("reference") or ""),
                    "import_profile_id": str(profile.get("id") or ""),
                    "import_profile_name": str(profile.get("name") or ""),
                }
            )

    for summary in scopes.values():
        summary["quality_rate"] = _quality_rate(summary["valid_rows"], summary["total_rows"])
    aggregate["quality_rate"] = _quality_rate(aggregate["valid_rows"], aggregate["total_rows"])
    return {
        "project_id": project_id,
        "last_imported_at": last_imported_at,
        "summary": aggregate,
        "scope_summary": list(scopes.values()),
        "batches": batches,
    }


def build_import_error_csv(*, batch_id: str) -> tuple[bytes, str]:
    """Build an editable error-only CSV from a persisted import preflight batch."""
    batch = get_import_batch(batch_id)
    if batch is None:
        raise ValueError("未找到导入预检批次")

    error_rows = [row for row in batch["rows"] if row.get("error_message")]
    source_headers: list[str] = []
    for row in error_rows:
        for key in (row.get("source_row") or {}).keys():
            if key not in source_headers:
                source_headers.append(key)

    output = StringIO(newline="")
    fieldnames = ["batch_id", "scope", "row_number", "preflight_action", "error_message", *source_headers]
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in error_rows:
        writer.writerow(
            {
                "batch_id": batch["id"],
                "scope": batch["scope"],
                "row_number": row["row_number"],
                "preflight_action": row["action"],
                "error_message": row["error_message"],
                **(row.get("source_row") or {}),
            }
        )
    filename = f"import_errors_{batch['scope']}_{batch['id'][:8]}.csv"
    return output.getvalue().encode("utf-8-sig"), filename


def _apply_operation(
    operation: dict[str, Any],
    *,
    actor_name: str,
    file_name: str,
    row_number: int,
    source: dict[str, str] | None = None,
) -> None:
    action = operation["action"]
    entity_type = operation["entity_type"]
    payload = operation["operation"]
    detail: dict[str, Any] = {"file_name": file_name, "row": row_number, "batch_import": True}
    if source:
        detail["external_readonly_source"] = source
    if entity_type == "project":
        if action == "create":
            project = create_project(**payload)
        else:
            project = update_project(payload["project_id"], {key: value for key, value in payload.items() if key != "project_id"})
        if project is None:
            raise ValueError("预检对应项目已不存在")
        _audit(project["id"], "project", project["id"], actor_name, detail)
        return
    if entity_type == "site":
        if action == "create":
            site = create_site(**payload)
        else:
            site_id = payload["site_id"]
            site = update_site(site_id, {key: value for key, value in payload.items() if key != "site_id"})
        if site is None:
            raise ValueError("预检对应中心已不存在")
        _audit(site["project_id"], "site", site["id"], actor_name, detail)
        return
    site_id = payload["site_id"]
    save_subject_codes(site_id, [{"code": payload["subject_code"], "enrollment_status": payload["enrollment_status"]}])
    _audit(payload["project_id"], "subject_code", f"{site_id}:{payload['subject_code']}", actor_name, detail)


def commit_master_data_import(*, batch_id: str, actor_name: str = "项目管理员") -> dict[str, Any]:
    current = get_import_batch(batch_id)
    if current is None:
        raise ValueError("未找到导入预检批次")
    if current["status"] == "committed":
        return current
    summary = {"total": len(current["rows"]), "created": 0, "updated": 0, "skipped": 0}
    errors: list[dict[str, Any]] = []
    source = {
        "system": str(current.get("source_system") or "").strip(),
        "reference": str(current.get("source_reference") or "").strip(),
        "exported_at": str(current.get("source_exported_at") or "").strip(),
    }
    source = {key: value for key, value in source.items() if value}
    for row in current["rows"]:
        if row.get("error_message") or row.get("action") not in {"create", "update"}:
            summary["skipped"] += 1
            if row.get("error_message"):
                errors.append({"row": row["row_number"], "message": row["error_message"]})
            continue
        try:
            _apply_operation(
                row,
                actor_name=actor_name,
                file_name=current["file_name"],
                row_number=row["row_number"],
                source=source or None,
            )
            summary["created" if row["action"] == "create" else "updated"] += 1
        except ValueError as exc:
            summary["skipped"] += 1
            errors.append({"row": row["row_number"], "message": str(exc)})
    with transaction() as connection:
        connection.execute(
            """
            UPDATE import_batches
            SET status = 'committed', committed_summary_json = ?, committed_at = ?
            WHERE id = ?
            """,
            (json.dumps({**summary, "errors": errors}, ensure_ascii=False), _now(), batch_id),
        )
    return get_import_batch(batch_id) or {}


def import_master_data(
    *,
    scope: str,
    file_name: str,
    content: bytes,
    default_project_id: str = "",
    default_site_id: str = "",
    actor_name: str = "演示管理员",
) -> dict[str, Any]:
    preview = preview_master_data_import(
        scope=scope,
        file_name=file_name,
        content=content,
        default_project_id=default_project_id,
        default_site_id=default_site_id,
        actor_name=actor_name,
    )
    committed = commit_master_data_import(batch_id=preview["id"], actor_name=actor_name)
    summary = dict(committed.get("committed_summary") or {})
    return {
        "scope": committed["scope"],
        "file_name": committed["file_name"],
        "summary": {key: int(summary.get(key, 0)) for key in ("total", "created", "updated", "skipped")},
        "errors": list(summary.get("errors") or []),
    }
